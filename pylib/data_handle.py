# -*- coding: utf-8 -*-
'''
    @File    CSV Excel handle  [python3.5]
    @Author  tx
    @Created On 2018-04-26
    @Updated On 2018-04-28

    据说：
    2007版以前的Excel（xls结尾的），需要使用xlrd读，xlwt写。
    2007版以后的Excel（xlsx结尾的），需要使用openpyxl来读写。
    from pyExcelerator import *
'''
import os
import xlrd
import xlwt
import openpyxl
import csv
import json


def get_outpath(path1='output', path2='test'):
    pwd = os.path.dirname(os.path.realpath(__file__))           #pwd2 = sys.path[0]
    # print(pwd)
    pardir = os.path.abspath(os.path.join(pwd, os.pardir))
    dpath =  os.path.join(pardir, path1, path2)
    try:
        if not os.path.exists(dpath):
            os.makedirs(dpath)
        return dpath
    except Exception as e:
        print(e)
        return False


def save_data(buf, fpath, fname, ftype=None):
    try:
        if not os.path.exists(fpath):
            os.makedirs(fpath)
        fullpath = os.path.join(fpath, fname)
        if ftype == 'json':
            with open(fullpath, 'w') as fjp:
                json.dump(buf, fjp)
        else:
            with open(fullpath, 'w') as fp:
                # fp.write(buf.encode('utf-8', 'ignore'))
                fp.write(buf.encode('gbk', 'ignore'))
        return True
    except Exception as e:
        print(e)
        return False


def get_data(fpath, fname, ftype=None):
    try:
        fullpath = os.path.join(fpath, fname)
        if ftype == 'json':
            with open(fullpath, 'r') as fjp:
                buf = json.load(fjp)
        else:
            with open(fullpath, 'r') as fp:
                buf = fp.read()
        return buf
    except Exception as e:
        print(e)
        return None


class CSVHandle(object):
    def __init__(self):
        pass

    def writer_csv(self, fname, data):
        with open(fname, 'w', newline='', encoding='utf-8-sig') as fp:
            writer = csv.writer(fp)
            writer.writerows(data)

    def read_csv(self, fname):
        results = []
        with open(fname, encoding="utf-8") as fp:
            reader = csv.reader(fp, delimiter=',')
            header = next(reader)                           # 获取表头
            print('header: ', header)
            results = [row for row in reader]
            return results

    def dic_write_csv(self, fname, dic_list):
        header = [k for k in dic_list[0]]                   # ['age', 'name']
        with open(fname, 'w') as fp:
            writer = csv.DictWriter(fp, delimiter=',', fieldnames=header)
            writer.writeheader()
            for dit in dic_list:
                print('write value: ', dit)
                writer.writerow(dit)

    def dic_read_csv(self, fname):
        results = []
        with open(fname, 'r') as fp:
            reader = csv.DictReader(fp)
            header = reader.fieldnames                      # 获取表头
            print('header: ', header)
            for i in range(1):                              # 忽略第一行数据
                ignore = next(reader)
                # print(ignore)
            results = [row for row in reader]
            return results


class ExcelHandle(object):
    def __init__(self):
        pass

    def read_excel(self, fname):
        book = xlrd.open_workbook(fname)
        sheets = book.sheet_names()                     # 获取所有表名
        sheet = book.sheet_by_name(sheets[0])
        nrows = sheet.nrows                             # sheet.nrows: 行数
        ncols = sheet.ncols                             # sheet.ncols: 列数
        for i in range(0, nrows):
            # row = sheet.row(i)
            # rowValue = sheet.row_values(i)
            for j in range(0, ncols):
                print(sheet.cell_value(i, j), '\t', end='')
            print()

    def write_excel(self, fname, sheetName, values):
        book  = xlwt.Workbook(encoding='utf-8', style_compression=0)
        sheet = book.add_sheet(sheetName, cell_overwrite_ok=True)

        for row in range(0, len(values)):
            for column in range(0, len(values[row])):
                sheet.write(row, column, values[row][column])
        book.save(fname)

    def read07_excel(self, fname, debug=0):
        book = openpyxl.load_workbook(fname)
        sheets = book.sheetnames                        # 获取所有表名
        sheet = book[sheets[0]]                         # sheet = book.get_sheet_by_name(sname)
        nrows = sheet.rows                              # type: generator , # ncols = sheet.columns
        m_row = sheet.max_row                           # 最大的行数, sheet.max_column: 最大列数
        print(m_row)
        if not debug:
            results = []
            for row in nrows:
                c_list = [str(j.value) for j in row if j.value]
                results.append(c_list)
            return results
        else:
        # debug 状态
            print(sheet['A1'].value)                    # 获取单元格A1值, sheet['A1'].column:列值, sheet['A1'].row:行号
            print(sheet.cell(row=1,column=1).value)     # 获取单元格A1值, (注: 下标是从1 开始的)
            for row in nrows:
                for j in row:
                    print(j.value, '\t', end='')
                print()

    def write07_excel(self, fname, sname, values):
        book = openpyxl.Workbook()
        book.save(fname)
        print ("新建Excel："+fname+"成功")

        book = openpyxl.load_workbook(fname)
        sheet = book.active
        sheet.title = sname

        fields = values[0]
        for field in range(1, len(fields)+1):
            sheet.cell(row=1, column=field, value=str(fields[field-1]))

        for row in range(0, len(values)-1):
            for column in range(0, len(values[row])):
                sheet.cell(row=row+2, column=column+1, value=str(values[row+1][column]))

        # for row in range(0, len(values)):
        #     for column in range(0, len(values[row])):
        #         sheet.cell(row=row+1, column=column+1, value=str(values[row][column]))

        book.save(fname)



def main():
    xlObj = ExcelHandle()
    csvObj = CSVHandle()



if __name__ == '__main__':
    main()